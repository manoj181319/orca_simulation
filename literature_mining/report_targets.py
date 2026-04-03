"""
report_targets.py  —  Phase 0, Step 4
Generates a formatted Excel report for Siva Sir with four sheets:
    1. About              — project metadata
    2. Redox Protein Targets — full ranked table; non-simulatable genes visually marked
    3. Ligand Mentions    — signal ligands + separate noise ligand section
    4. Recommended Pairs  — simulatable pairs with ligand basis, role, and docking status

Fixes applied in this version:
    - "Siva Sir" used consistently everywhere (was "Shiva")
    - Sheet 2 SIMULATABLE set updated to match score_targets.py (H2O2, organic
      hydroperoxide, HHQ, Mn2+, Fe2+, [2Fe-2S], SAM now correctly marked Yes)
    - Sheet 2 splits signal ligands and noise ligands into separate sections
    - Sheet 1 adds Simulatable and Ligand Note columns; non-simulatable genes
      are grey-striked so Siva Sir can immediately see which are excluded
    - Sheet 4 adds Ligand Role and Basis columns for traceability
    - soxR [2Fe-2S] gets a special note in the Notes column warning about
      metal-cluster handling in ORCA inputs
    - trxA NADPH gets a note clarifying NADPH acts via TrxB, not direct binding
    - Reference Paper column added to both Redox Protein Targets and Recommended
      Pairs sheets — clickable hyperlinks to PDB, PubMed, PMC, or UniProt

Input:
    literature_mining/extracted/scored_targets.json

Output:
    literature_mining/redox_amr_targets_Pa.xlsx

Usage:
    python report_targets.py
"""

import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
INPUT_FILE  = Path(__file__).parent / "extracted" / "scored_targets.json"
OUTPUT_FILE = Path(__file__).parent / "redox_amr_targets_Pa.xlsx"

UNIPROT_BASE = "https://www.uniprot.org/uniprotkb?query={gene}+pseudomonas+aeruginosa&reviewed=true"

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
HEADER_BG      = "1F4E79"
SUBHEADER_BG   = "2E75B6"
ALT_ROW_BG     = "DEEAF1"
HIGHLIGHT_BG   = "E2EFDA"
WARN_BG        = "FFF2CC"
EXCLUDED_BG    = "F2F2F2"   # grey for non-simulatable rows in sheet 1
NOISE_BG       = "FCE4D6"   # light orange for noise ligand section
SPECIAL_BG     = "EAF3DE"   # light green for special-note rows in sheet 4
BORDER_COLOR   = "BFBFBF"

# ---------------------------------------------------------------------------
# Simulatable ligands — must match SIMULATABLE_LIGANDS in score_targets.py
# ---------------------------------------------------------------------------
SIMULATABLE = {
    "NADH", "NADPH", "FAD", "FMN", "heme", "haem", "ubiquinone", "menaquinone",
    "glutathione", "ferredoxin", "CoA", "coenzyme A", "flavin", "quinone", "SAM",
    "3OC12-HSL", "3-oxo-C12-HSL", "C4-HSL", "N-butanoyl-HSL",
    "H2O2", "hydrogen peroxide", "organic hydroperoxide",
    "Fe2+", "Mn2+", "[2Fe-2S]", "HHQ", "PQS",
}

# Noise ligands — globally prevalent in Pa literature, excluded from signal scoring
LIGAND_NOISE = {"pyocyanin", "phenazine", "iron", "pyoverdine", "pyochelin"}

# Special handling notes appended to specific pairs in sheet 4
SPECIAL_NOTES = {
    "soxR":  "[2Fe-2S] is a metal cluster, NOT a standard organic ligand. "
             "Requires special ORCA input treatment: add %coords block with Fe-S "
             "cluster geometry and set correct spin multiplicity. Confirm approach "
             "with Siva Sir before building input file.",
    "trxA":  "NADPH acts via TrxB (thioredoxin reductase), not by direct binding "
             "to TrxA. The simulation target for this pair is the TrxB-NADPH "
             "complex. Consider using trxB+FAD instead if direct binding is required.",
    "sodA":  "Mn2+ is the active-site metal ion cofactor, not a free ligand. "
             "Use the metal-cofactored form of the protein from PDB for simulation.",
    "sodB":  "Fe2+ is the active-site metal ion cofactor, not a free ligand. "
             "Use the metal-cofactored form of the protein from PDB for simulation.",
    "sodM":  "Mn2+ is the active-site metal ion cofactor. Cambialistic — can use "
             "Fe2+ under iron-replete conditions. Confirm metal state with Siva Sir.",
    "fur":   "Fe2+ is the corepressor ion that triggers DNA binding. Structural "
             "Zn2+ is also present but is not the regulatory cofactor.",
    "ahpC":  "Organic hydroperoxide is the substrate, not a stable co-crystal "
             "ligand. Use a specific hydroperoxide (e.g. cumene hydroperoxide or "
             "tert-butyl hydroperoxide) for docking — confirm choice with Siva Sir.",
    "mvfR":  "HHQ (2-heptyl-4-hydroxyquinoline) is the confirmed PqsR binding "
             "ligand. PQS (Pseudomonas quinolone signal) also binds PqsR. "
             "Both are valid — confirm which to simulate with Siva Sir.",
}


# ---------------------------------------------------------------------------
# GENE_REFERENCES — authoritative reference URLs for every gene.
#
# Used as the primary lookup for the "Reference Paper" column in both sheets.
# Format: { gene: (url, display_text) }
# Hierarchy: PDB co-crystal > primary literature (PMC/PubMed) > UniProt search.
# Falls back to basis_to_url() for any gene not listed here.
# ---------------------------------------------------------------------------
GENE_REFERENCES = {
    # Catalases
    "katA":  ("https://www.rcsb.org/structure/4E37",
              "PDB 4E37 — KatA + heme + NADPH"),
    "katB":  ("https://pmc.ncbi.nlm.nih.gov/articles/PMC177506/",
              "PMC177506 — KatB heme catalase in Pa"),
    "katE":  ("https://www.uniprot.org/uniprotkb?query=katE+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — KatE Pa"),
    # Peroxiredoxin system
    "ahpC":  ("https://www.sciencedirect.com/science/article/pii/S2213231721002342",
              "ScienceDirect — AhpC neutralizes hydroperoxides"),
    "ahpF":  ("https://www.sciencedirect.com/science/article/pii/S2213231721002342",
              "ScienceDirect — AhpCF peroxidase system"),
    # Redox-sensing regulators
    "oxyR":  ("https://www.sciencedirect.com/science/article/pii/S0923250811001756",
              "ScienceDirect — OxyR H2O2 sensor Pa"),
    "soxR":  ("https://pubmed.ncbi.nlm.nih.gov/9614965/",
              "PubMed 9614965 — SoxR in Pa"),
    "soxS":  ("https://pubmed.ncbi.nlm.nih.gov/9614965/",
              "PubMed 9614965 — SoxRS system Pa"),
    "mexR":  ("https://pubmed.ncbi.nlm.nih.gov/34912548/",
              "PubMed 34912548 — MexR/NalD redox sensing"),
    # Superoxide dismutases
    "sodA":  ("https://pmc.ncbi.nlm.nih.gov/articles/PMC206923/",
              "PMC206923 — sodA cloning and characterisation Pa"),
    "sodB":  ("https://pmc.ncbi.nlm.nih.gov/articles/PMC177481/",
              "PMC177481 — sodB FeSOD in Pa"),
    "sodM":  ("https://pmc.ncbi.nlm.nih.gov/articles/PMC1828791/",
              "PMC1828791 — sodM cambialistic SOD Pa virulence"),
    # Thioredoxin system
    "trxA":  ("https://pubmed.ncbi.nlm.nih.gov/22609922/",
              "PubMed 22609922 — Pa thioredoxin-dependent peroxidase"),
    "trxB":  ("https://en.wikipedia.org/wiki/Thioredoxin_reductase",
              "Wikipedia — Thioredoxin reductase FAD cofactor"),
    "trxC":  ("https://www.uniprot.org/uniprotkb?query=trxC+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — TrxC Pa"),
    # Glutaredoxin / glutathione system
    "grxA":  ("https://www.uniprot.org/uniprotkb?query=grxA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — GrxA Pa"),
    "grxB":  ("https://www.uniprot.org/uniprotkb?query=grxB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — GrxB Pa"),
    "gshA":  ("https://www.uniprot.org/uniprotkb?query=gshA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — GshA Pa"),
    "gshB":  ("https://www.uniprot.org/uniprotkb?query=gshB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — GshB Pa"),
    "gor":   ("https://www.uniprot.org/uniprotkb?query=gor+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — Glutathione reductase Pa"),
    # NADH/NADPH metabolism & electron transport
    "ndh":   ("https://www.uniprot.org/uniprotkb?query=ndh+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — NADH dehydrogenase II Pa"),
    "nuoA":  ("https://www.uniprot.org/uniprotkb?query=nuoA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — NuoA Complex I Pa"),
    "nuoB":  ("https://www.uniprot.org/uniprotkb?query=nuoB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — NuoB Complex I Pa"),
    "nuoN":  ("https://www.uniprot.org/uniprotkb?query=nuoN+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — NuoN Complex I Pa"),
    "sdhA":  ("https://www.uniprot.org/uniprotkb?query=sdhA+pseudomonas+aeruginosa+FAD&reviewed=true",
              "UniProt — SdhA Complex II FAD-binding subunit"),
    "sdhB":  ("https://www.uniprot.org/uniprotkb?query=sdhB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — SdhB Complex II Fe-S subunit"),
    # Cytochromes
    "cioA":  ("https://www.uniprot.org/uniprotkb?query=cioA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — CioA cyanide-insensitive oxidase"),
    "cioB":  ("https://www.uniprot.org/uniprotkb?query=cioB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — CioB"),
    "cyoA":  ("https://www.uniprot.org/uniprotkb?query=cyoA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — CyoA cytochrome bo3"),
    "coxA":  ("https://www.uniprot.org/uniprotkb?query=coxA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — CoxA cytochrome c oxidase"),
    # Ferredoxins
    "fdxA":  ("https://www.uniprot.org/uniprotkb?query=fdxA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — FdxA ferredoxin I Pa"),
    "fdxB":  ("https://www.uniprot.org/uniprotkb?query=fdxB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — FdxB ferredoxin II Pa"),
    # Iron metabolism
    "fur":   ("https://pmc.ncbi.nlm.nih.gov/articles/PMC5648857/",
              "PMC5648857 — Fur Fe2+ corepressor Pa"),
    "pvdA":  ("https://pubmed.ncbi.nlm.nih.gov/17015659/",
              "PubMed 17015659 — PvdA ornithine hydroxylase Pa"),
    "pvdD":  ("https://www.uniprot.org/uniprotkb?query=pvdD+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — PvdD NRPS"),
    "pchA":  ("https://www.uniprot.org/uniprotkb?query=pchA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — PchA pyochelin biosynthesis"),
    "bfrA":  ("https://www.uniprot.org/uniprotkb?query=bfrA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — BfrA bacterioferritin Pa"),
    "bfrB":  ("https://www.uniprot.org/uniprotkb?query=bfrB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — BfrB bacterioferritin Pa"),
    # Phenazine biosynthesis
    "phzA1": ("https://www.uniprot.org/uniprotkb?query=phzA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — PhzA1 Pa"),
    "phzB1": ("https://www.uniprot.org/uniprotkb?query=phzB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — PhzB1 Pa"),
    "phzM":  ("https://pubmed.ncbi.nlm.nih.gov/17376257/",
              "PubMed 17376257 — PhzM SAM methyltransferase Pa"),
    "phzS":  ("https://pubmed.ncbi.nlm.nih.gov/17376257/",
              "PubMed 17376257 — PhzS flavin hydroxylase Pa"),
    "phzH":  ("https://www.uniprot.org/uniprotkb?query=phzH+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — PhzH Pa"),
    # Efflux pumps
    "mexA":  ("https://www.uniprot.org/uniprotkb?query=mexA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexA efflux fusion protein"),
    "mexB":  ("https://www.uniprot.org/uniprotkb?query=mexB+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexB RND pump"),
    "oprM":  ("https://www.uniprot.org/uniprotkb?query=oprM+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — OprM outer membrane channel"),
    "mexC":  ("https://www.uniprot.org/uniprotkb?query=mexC+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexC"),
    "mexD":  ("https://www.uniprot.org/uniprotkb?query=mexD+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexD"),
    "mexE":  ("https://www.uniprot.org/uniprotkb?query=mexE+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexE"),
    "mexF":  ("https://www.uniprot.org/uniprotkb?query=mexF+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexF"),
    "mexX":  ("https://www.uniprot.org/uniprotkb?query=mexX+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexX"),
    "mexY":  ("https://www.uniprot.org/uniprotkb?query=mexY+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — MexY"),
    # Quorum sensing
    "lasR":  ("https://www.rcsb.org/structure/3IX3",
              "PDB 3IX3 — LasR + 3OC12-HSL co-crystal"),
    "lasI":  ("https://pubmed.ncbi.nlm.nih.gov/15306017/",
              "PubMed 15306017 — LasI 3OC12-HSL synthase"),
    "rhlR":  ("https://www.rcsb.org/structure/8B4A",
              "PDB 8B4A — RhlR + C4-HSL co-crystal"),
    "rhlI":  ("https://www.nature.com/articles/s41467-023-43702-4",
              "Nature Commun. — RhlI C4-HSL synthase"),
    "mvfR":  ("https://www.nature.com/articles/s41467-023-43702-4",
              "Nature Commun. — PqsR binds HHQ/PQS"),
    "pqsA":  ("https://www.uniprot.org/uniprotkb?query=pqsA+pseudomonas+aeruginosa&reviewed=true",
              "UniProt — PqsA anthraniloyl-CoA ligase"),
}


def get_gene_ref(gene: str, basis: str):
    """Return (url, display_text) for a gene's reference paper.
    Checks GENE_REFERENCES first; falls back to parsing basis_to_url()."""
    if gene in GENE_REFERENCES:
        return GENE_REFERENCES[gene]
    return basis_to_url(basis)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def hf(size=11, color="FFFFFF", bold=True):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def bf(size=10, color="000000", italic=False):
    return Font(name="Calibri", size=size, color=color, italic=italic)

def fill(color):
    return PatternFill("solid", fgColor=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def header_row(ws, row, values, bg=HEADER_BG):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font  = hf()
        c.fill  = fill(bg)
        c.alignment = center()
        c.border = border()

def body_row(ws, row, values, bg=None, italic=False, font_color="000000"):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font  = bf(color=font_color, italic=italic)
        if bg:
            c.fill = fill(bg)
        c.alignment = left()
        c.border = border()

def hyperlink_cell(ws, row, col, url, text):
    c = ws.cell(row=row, column=col, value=text)
    c.hyperlink = url
    c.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    c.alignment = left()
    c.border = border()

def freeze_filter(ws, cell="A2"):
    ws.freeze_panes = cell
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Reference URL parser
#
# Parses the ligand_basis string (from GENE_BIOCHEMISTRY) into a clickable
# URL. Returns (url, display_text). If no parseable identifier is found,
# returns (None, basis_text) so the cell renders as plain text.
#
# Patterns handled:
#   PDB XXXX        → https://www.rcsb.org/structure/XXXX (first ID if multiple)
#   PMC\d+          → https://pmc.ncbi.nlm.nih.gov/articles/PMCXXXXXX/
#   PubMed \d+      → https://pubmed.ncbi.nlm.nih.gov/XXXXXXXX/
#   ScienceDirect S…→ https://www.sciencedirect.com/science/article/pii/SXXXXXXXX
#   Nature Comm s…  → https://www.nature.com/articles/sXXXXX
#   OUP \d+         → https://academic.oup.com/nar/article/XXXXXXXX
#   Wikipedia XXX   → https://en.wikipedia.org/wiki/XXX
# ---------------------------------------------------------------------------
def basis_to_url(basis: str):
    if not basis:
        return None, ""

    # PDB IDs — 4-character alphanumeric e.g. "PDB 4E37" or "PDB 2UV0, 3IX3"
    pdb_match = re.search(r'\bPDB\s+([A-Z0-9]{4})', basis, re.IGNORECASE)
    if pdb_match:
        pdb_id = pdb_match.group(1).upper()
        return f"https://www.rcsb.org/structure/{pdb_id}", f"PDB {pdb_id}"

    # PMC IDs — "PMC177506"
    pmc_match = re.search(r'\bPMC(\d+)\b', basis, re.IGNORECASE)
    if pmc_match:
        pmc_id = pmc_match.group(1)
        return f"https://pmc.ncbi.nlm.nih.gov/articles/PMC{pmc_id}/", f"PMC{pmc_id}"

    # PubMed IDs — "PubMed 15306017"
    pubmed_match = re.search(r'\bPubMed\s+(\d+)\b', basis, re.IGNORECASE)
    if pubmed_match:
        pmid = pubmed_match.group(1)
        return f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/", f"PubMed {pmid}"

    # ScienceDirect PII — "ScienceDirect S0923250811001756"
    sd_match = re.search(r'\bScienceDirect\s+(S[0-9A-Z]+)\b', basis, re.IGNORECASE)
    if sd_match:
        pii = sd_match.group(1)
        return f"https://www.sciencedirect.com/science/article/pii/{pii}", f"ScienceDirect"

    # Nature Communications DOI slug — "Nature Comm s41467-023-43702-4"
    nat_match = re.search(r'\bNature Comm\w*\s+(s[\d\-]+)\b', basis, re.IGNORECASE)
    if nat_match:
        slug = nat_match.group(1)
        return f"https://www.nature.com/articles/{slug}", f"Nature Commun."

    # OUP article ID — "OUP 8222440"
    oup_match = re.search(r'\bOUP\s+(\d+)\b', basis, re.IGNORECASE)
    if oup_match:
        article_id = oup_match.group(1)
        return f"https://academic.oup.com/nar/article/{article_id}", "OUP / NAR"

    # Wikipedia article — "Wikipedia Thioredoxin_reductase"
    wiki_match = re.search(r'\bWikipedia\s+(\S+)', basis, re.IGNORECASE)
    if wiki_match:
        title = wiki_match.group(1).replace(" ", "_")
        return f"https://en.wikipedia.org/wiki/{title}", f"Wikipedia"

    # No parseable identifier — return plain text
    return None, basis[:60]


# ---------------------------------------------------------------------------
# Sheet 1 — About
# ---------------------------------------------------------------------------
def build_cover(wb):
    ws = wb["About"]
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 16, "B": 72})

    ws.row_dimensions[1].height = 36
    title = ws.cell(row=1, column=1, value="Redox AMR Targets — Pseudomonas aeruginosa")
    title.font = Font(name="Calibri", bold=True, size=14, color=HEADER_BG)
    title.alignment = left()
    ws.merge_cells("A1:B1")

    meta = [
        ("Project",      "UV-Induced Excitation in Bacterial Proteins"),
        ("Organism",     "Pseudomonas aeruginosa PAO1"),
        ("Focus",        "Redox-linked antimicrobial resistance"),
        ("Phase",        "Phase 0 — Literature Mining"),
        ("Coordinator",  "Siva Sir"),
        ("Generated by", "orca_simulation pipeline (report_targets.py)"),
        ("Description",
            "This report identifies protein and gene candidates from P. aeruginosa "
            "involved in redox biology and antimicrobial resistance. Candidates are "
            "ranked by citation frequency, PDB structure availability, and co-mention "
            "with simulatable ligands. Ligand assignments are biochemistry-derived "
            "(GENE_BIOCHEMISTRY table in extract_targets.py) — NOT corpus statistics. "
            "Non-simulatable genes (passive channels, structural subunits) are marked "
            "in grey on the Redox Protein Targets sheet and excluded from Recommended Pairs."
        ),
        ("Ligand noise",
            "Pyocyanin, phenazine, iron, pyoverdine, and pyochelin are globally "
            "prevalent in P. aeruginosa literature. They have been separated into a "
            "noise section on the Ligand Mentions sheet and are excluded from ligand "
            "diversity scoring."
        ),
    ]

    for row, (key, val) in enumerate(meta, 2):
        k = ws.cell(row=row, column=1, value=key)
        k.font = bf(bold=True) if False else Font(name="Calibri", bold=True, size=11)
        k.alignment = left()
        v = ws.cell(row=row, column=2, value=val)
        v.font = Font(name="Calibri", size=11)
        v.alignment = left(wrap=True)
        ws.row_dimensions[row].height = 22 if row != 8 else 60


# ---------------------------------------------------------------------------
# Sheet 2 — Redox Protein Targets
# ---------------------------------------------------------------------------
def build_sheet1(wb, ranked_genes):
    ws = wb.create_sheet("Redox Protein Targets")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    headers = [
        "Rank", "Gene", "Protein Name", "Function", "Category",
        "Papers", "Score", "Has PDB", "Top Ligand", "Simulatable",
        "Ligand Note", "Reference Paper", "UniProt Search"
    ]
    header_row(ws, 1, headers)

    for i, gd in enumerate(ranked_genes, 2):
        score       = gd["final_score"]
        simulatable = gd.get("simulatable", False)
        ligand      = gd.get("top_ligand", "")
        note        = gd.get("ligand_note", "")
        basis       = gd.get("ligand_basis", "")

        if not simulatable:
            bg         = EXCLUDED_BG
            font_color = "888888"
            italic     = True
        elif score >= 0.70:
            bg         = HIGHLIGHT_BG
            font_color = "000000"
            italic     = False
        elif i % 2 == 0:
            bg         = ALT_ROW_BG
            font_color = "000000"
            italic     = False
        else:
            bg         = None
            font_color = "000000"
            italic     = False

        row_values = [
            gd["rank"],
            gd["gene"],
            gd["protein"],
            gd["function"],
            gd["category"].replace("_", " ").title(),
            gd["paper_count"],
            round(score, 4),
            "Yes" if gd["has_pdb"] else "No",
            ligand if ligand != "no_ligand" else "—",
            "Yes" if simulatable else "No",
            note,
            "",   # Reference Paper — hyperlink added separately below
            "",   # UniProt Search  — hyperlink added separately below
        ]
        body_row(ws, i, row_values, bg=bg, italic=italic, font_color=font_color)

        # Column 12 — Reference Paper (clickable link to PDB / PMC / PubMed etc.)
        ref_url, ref_text = get_gene_ref(gd["gene"], basis)
        if ref_url:
            hyperlink_cell(ws, i, 12, ref_url, ref_text)
        else:
            c = ws.cell(row=i, column=12, value=ref_text)
            c.font = bf(color=font_color, italic=italic)
            if bg:
                c.fill = fill(bg)
            c.alignment = left()
            c.border = border()

        # Column 13 — UniProt Search
        hyperlink_cell(ws, i, 13,
                       UNIPROT_BASE.format(gene=gd["gene"]),
                       f"Search {gd['gene']}")
        ws.row_dimensions[i].height = 22

    set_widths(ws, {
        "A": 6,  "B": 9,  "C": 32, "D": 45, "E": 18,
        "F": 8,  "G": 8,  "H": 9,  "I": 20, "J": 12,
        "K": 40, "L": 20, "M": 18,
    })
    freeze_filter(ws)


# ---------------------------------------------------------------------------
# Sheet 3 — Ligand Mentions
# ---------------------------------------------------------------------------
def build_sheet2(wb, scored_data):
    ws = wb.create_sheet("Ligand Mentions")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    # Aggregate signal and noise ligand counts separately
    signal_totals = {}
    noise_totals  = {}

    for gd in scored_data.get("ranked_genes", []):
        for lig, cnt in gd.get("ligands_co_mentioned", {}).items():
            signal_totals[lig] = signal_totals.get(lig, 0) + cnt
        for lig, cnt in gd.get("ligands_noise_co_mentioned", {}).items():
            noise_totals[lig] = noise_totals.get(lig, 0) + cnt

    sorted_signal = sorted(signal_totals.items(), key=lambda x: -x[1])
    sorted_noise  = sorted(noise_totals.items(),  key=lambda x: -x[1])

    headers = ["Rank", "Ligand / Cofactor", "Co-mention Count", "Simulatable?", "Notes"]
    header_row(ws, 1, headers, bg=SUBHEADER_BG)

    row = 2
    # Signal ligands
    for rank, (lig, cnt) in enumerate(sorted_signal, 1):
        sim = "Yes" if lig in SIMULATABLE else "No"
        bg  = HIGHLIGHT_BG if sim == "Yes" else (ALT_ROW_BG if rank % 2 == 0 else None)
        note = ""
        if lig == "[2Fe-2S]":
            note = "Metal cluster — requires special ORCA input. See notes in Recommended Pairs."
        body_row(ws, row, [rank, lig, cnt, sim, note], bg=bg)
        ws.row_dimensions[row].height = 20
        row += 1

    # Section divider for noise ligands
    row += 1
    divider_cell = ws.cell(row=row, column=1,
        value="NOISE LIGANDS — globally prevalent in Pa literature; excluded from scoring")
    divider_cell.font = Font(name="Calibri", bold=True, size=10, color="9C3006")
    divider_cell.fill = fill(NOISE_BG)
    divider_cell.alignment = left()
    divider_cell.border = border()
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 22
    row += 1

    # Noise ligand sub-header
    header_row(ws, row, headers, bg="C55A11")
    ws.row_dimensions[row].height = 22
    row += 1

    noise_note = (
        "Appears in >90% of Pa redox papers. Excluded from ligand assignment — "
        "co-occurrence reflects shared biological context, not direct binding."
    )
    for rank, (lig, cnt) in enumerate(sorted_noise, 1):
        body_row(ws, row, [rank, lig, cnt, "No (noise)", noise_note], bg=NOISE_BG)
        ws.row_dimensions[row].height = 20
        row += 1

    set_widths(ws, {"A": 6, "B": 24, "C": 18, "D": 14, "E": 60})
    freeze_filter(ws)


# ---------------------------------------------------------------------------
# Sheet 4 — Recommended Pairs
# ---------------------------------------------------------------------------
def build_sheet3(wb, recommended_pairs):
    ws = wb.create_sheet("Recommended Pairs")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    headers = [
        "Priority", "Gene", "Protein Name", "Suggested Ligand",
        "Ligand Role", "Score", "Has PDB Structure",
        "Needs Docking?", "Ligand Basis", "Reference Paper", "Notes"
    ]
    header_row(ws, 1, headers, bg=SUBHEADER_BG)

    for i, pair in enumerate(recommended_pairs, 2):
        gene    = pair["gene"]
        ligand  = pair["suggested_ligand"]
        docking = pair["needs_docking"]
        basis   = pair.get("ligand_basis", "")

        docking_text = "YES — request from Siva Sir" if docking else "No — pull from PDB"

        # Special notes: from SPECIAL_NOTES map; also flag [2Fe-2S]
        note = SPECIAL_NOTES.get(gene, "")
        if ligand == "[2Fe-2S]" and not note:
            note = ("[2Fe-2S] is a metal cluster. Requires special ORCA input treatment "
                    "— confirm approach with Siva Sir.")

        # Highlight rows with special notes
        if note:
            bg = SPECIAL_BG
        elif docking:
            bg = WARN_BG
        elif i % 2 == 0:
            bg = ALT_ROW_BG
        else:
            bg = None

        row_values = [
            pair["rank"],
            gene,
            pair["protein"],
            ligand if ligand != "no_ligand" else "—",
            pair.get("ligand_role", "").replace("_", " "),
            round(pair["score"], 4),
            "Yes" if pair["has_pdb"] else "No",
            docking_text,
            basis,
            "",   # Reference Paper — hyperlink added separately below
            note,
        ]
        body_row(ws, i, row_values, bg=bg)

        # Column 10 — Reference Paper clickable hyperlink
        ref_url, ref_text = get_gene_ref(gene, basis)
        if ref_url:
            hyperlink_cell(ws, i, 10, ref_url, ref_text)
        else:
            c = ws.cell(row=i, column=10, value=ref_text)
            c.font = bf()
            if bg:
                c.fill = fill(bg)
            c.alignment = left()
            c.border = border()

        ws.row_dimensions[i].height = 22

    set_widths(ws, {
        "A": 8,  "B": 9,  "C": 32, "D": 22, "E": 14,
        "F": 8,  "G": 18, "H": 28, "I": 55, "J": 32, "K": 65,
    })
    freeze_filter(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    if not INPUT_FILE.exists():
        print(f"ERROR: {INPUT_FILE} not found. Run score_targets.py first.")
        return

    with open(INPUT_FILE, encoding="utf-8") as f:
        scored_data = json.load(f)

    ranked_genes      = scored_data.get("ranked_genes", [])
    recommended_pairs = scored_data.get("recommended_pairs", [])

    if not ranked_genes:
        print("No ranked genes found in scored_targets.json. Exiting.")
        return

    print(f"Building Excel report for {len(ranked_genes)} genes "
          f"and {len(recommended_pairs)} recommended pairs...")

    wb = openpyxl.Workbook()
    wb.active.title = "About"
    build_cover(wb)
    build_sheet1(wb, ranked_genes)
    build_sheet2(wb, scored_data)
    build_sheet3(wb, recommended_pairs)
    wb.active = wb["Redox Protein Targets"]
    wb.save(OUTPUT_FILE)

    sim_count  = sum(1 for p in recommended_pairs)
    dock_count = sum(1 for p in recommended_pairs if p["needs_docking"])
    no_dock    = sim_count - dock_count

    print(f"\nReport saved to: {OUTPUT_FILE}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"\nRecommended pairs summary:")
    print(f"  Total simulatable pairs: {sim_count}")
    print(f"  Need docking (request Siva Sir): {dock_count}")
    print(f"  No docking (pull from PDB):      {no_dock}")
    print(f"\nNext step: send {OUTPUT_FILE.name} to Siva Sir for target confirmation.")


if __name__ == "__main__":
    main()
